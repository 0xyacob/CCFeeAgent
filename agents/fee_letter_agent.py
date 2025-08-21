#!/usr/bin/env python3
"""
Fee Letter Agent
===============

Modular agent for handling fee letter generation tasks.
This agent integrates with the existing fee letter automation system.
"""

import re
from typing import Dict, Any, Optional, List
from jinja2 import Template
from datetime import datetime  # unused
import json  # unused
import os

# Simple base classes to replace deleted imports
class BaseAgent:
    """Simple base agent class"""
    @property
    def name(self) -> str:
        return self.__class__.__name__

class TaskResult:
    """Simple task result class"""
    def __init__(self, success: bool, message: str, data: Dict[str, Any] = None):
        self.success = success
        self.message = message
        self.data = data or {}

# Import enhanced utilities
try:
    from vc_enhanced_utils import FuzzyMatcher, FeeCalculator, PolicyGate, ActivityLogger
except ImportError:
    FuzzyMatcher = None
    FeeCalculator = None
    PolicyGate = None
    ActivityLogger = None

# Import Microsoft Graph mail service
try:
    from microsoft_graph_mail import MicrosoftGraphMailService
except ImportError:
    MicrosoftGraphMailService = None


class FeeLetterAgent(BaseAgent):
    """
    Agent for handling fee letter generation and sending.
    
    This agent can:
    - Parse fee letter requests
    - Fetch investor and company data from Excel
    - Perform financial calculations
    - Generate personalized fee letters using Jinja2
    - Send emails via Microsoft Graph API
    - Track sent status
    """
    
    def __init__(self):
        # Configuration
        self.last_email_error: Optional[str] = None
        
        # Initialize activity logger
        self.activity_logger = ActivityLogger() if ActivityLogger else None
        
        # Class-level Excel adapter cache for performance across Streamlit reruns
        if not hasattr(FeeLetterAgent, "_excel_adapter_cache"):
            FeeLetterAgent._excel_adapter_cache = {}
        
        # Initialize Microsoft Graph mail service
        self.mail_service = None
        self.email_mode = os.getenv("EMAIL_MODE", "draft").lower()  # "draft" or "send"
        if MicrosoftGraphMailService:
            try:
                self.mail_service = MicrosoftGraphMailService(send_mode=self.email_mode)
                print(f"📧 Microsoft Graph mail service initialized (mode: {self.email_mode})")
            except Exception as e:
                print(f"⚠️ Could not initialize Microsoft Graph mail service: {e}")
        else:
            print("⚠️ Microsoft Graph mail service not available")
    
    @property
    def name(self) -> str:
        return "FeeLetterAgent"
    
    @property
    def patterns(self) -> List[str]:
        return [
            "create a fee letter for [Investor] for £[Amount] into [Company]",
            "create a letter for [Investor] to put £[Amount] into [Company]",
            "generate fee letter for [Investor] investing £[Amount] in [Company]",
            "[Investor] £[Amount] [Company]"
        ]
    
    def can_handle(self, prompt: str) -> bool:
        """Check if this agent can handle the fee letter prompt."""
        fee_patterns = [
            r"(?i)create a fee letter for (.+?) for £?([\d,]+\.?\d*) into (.+)",
            r"(?i)create a fee letter for (.+?) to put £?([\d,]+\.?\d*) into (.+)",
            r"(?i)create a letter for (.+?) for £?([\d,]+\.?\d*) into (.+)",
            r"(?i)create a letter for (.+?) to put £?([\d,]+\.?\d*) into (.+)",
            r"(?i)generate fee letter for (.+?) investing £?([\d,]+\.?\d*) in (.+)",
            r"(?i)(.+?) £?([\d,]+\.?\d*) (.+)"  # Simple format
        ]
        
        for pattern in fee_patterns:
            if re.match(pattern, prompt.strip()):
                return True
        return False
    
    def parse_request(self, prompt: str) -> dict:
        """
        Enhanced parser with robust intent + entity extraction
        Returns structured data with confidence scores and disambiguation
        """
        result = {
            'investor': {'name': None, 'confidence': 0.0, 'candidates': []},
            'company': {'name': None, 'confidence': 0.0, 'candidates': []},
            'amount': {'value': None, 'currency': 'GBP', 'confidence': 0.0},
            'investment_type': {'type': None, 'confidence': 0.0},  # net/gross
            'needs_clarification': False,
            'clarification_question': None,
            'extracted_entities': {},
            'confidence_threshold': 0.7
        }
        
        # Enhanced patterns for better extraction - prioritize full numbers FIRST
        amount_patterns = [
            # Prioritize full numbers like 50000 FIRST
            r'(?:£|GBP\s*)?(\d{4,})\s*(?:into|in|for)',
            r'(?:£|GBP\s*)?(\d{4,})',
            r'(?:£|GBP\s*)?(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s*(?:k|thousand)',
            r'(?:£|GBP\s*)?(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
            r'(\d+(?:,\d{3})*(?:\.\d{2})?)\s*(?:pounds?|gbp)',
        ]
        
        # Investment type patterns with scoring
        net_patterns = [r'\bnet\b', r'total\s+transfer', r'including\s+fees', r'all-in', r'net\s+investment']
        gross_patterns = [r'\bgross\b', r'before\s+fees', r'excluding\s+fees', r'investment\s+amount', r'gross\s+investment']
        
        # Extract investment amount
        for pattern in amount_patterns:
            match = re.search(pattern, prompt, re.IGNORECASE)
            if match:
                amount_str = match.group(1).replace(',', '')
                try:
                    amount = float(amount_str)
                    # Handle 'k' suffix
                    if 'k' in match.group(0).lower() or 'thousand' in match.group(0).lower():
                        amount *= 1000
                    result['amount']['value'] = amount
                    result['amount']['confidence'] = 0.9
                    break
                except ValueError:
                    continue
        
        # Extract investment type (net/gross) with confidence scoring
        net_score = sum(1 for pattern in net_patterns if re.search(pattern, prompt, re.IGNORECASE))
        gross_score = sum(1 for pattern in gross_patterns if re.search(pattern, prompt, re.IGNORECASE))
        
        if net_score > gross_score:
            result['investment_type']['type'] = 'net'
            result['investment_type']['confidence'] = min(0.95, 0.6 + (net_score * 0.15))
        elif gross_score > net_score:
            result['investment_type']['type'] = 'gross'
            result['investment_type']['confidence'] = min(0.95, 0.6 + (gross_score * 0.15))
        else:
            # Default to gross if ambiguous
            result['investment_type']['type'] = 'gross'
            result['investment_type']['confidence'] = 0.5
        
        # Enhanced name patterns with better context awareness
        name_patterns = [
            # Capture two-or-more word names after 'for' or 'to', but stop before amount/company cues
            r"(?i)(?:for|to)\s+([A-Za-z][A-Za-z'\-]+(?:\s+[A-Za-z][A-Za-z'\-]+)+?)(?=\s+(?:for|£|gbp|into|in|investing|wants?|with|on)\b)",
            r"(?i)investor[:\s]+([A-Za-z][A-Za-z'\-]+(?:\s+[A-Za-z][A-Za-z'\-]+)+)",
            r"(?i)^([A-Za-z][A-Za-z'\-]+\s+[A-Za-z][A-Za-z'\-]+)(?=\s)",
            r"(?i)letter\s+for\s+([A-Za-z][A-Za-z'\-]+\s+[A-Za-z][A-Za-z'\-]+)"
        ]
        
        # Company name patterns with better boundary detection
        company_patterns = [
            r'(?i)into\s+([A-Za-z0-9&\-\.\s]+?)(?:\s*$|,|\.|;)',
            r'(?i)company[:\s]+([A-Za-z0-9&\-\.\s]+?)(?:\s*$|,|\.|;)',
            r'(?i)investing\s+in\s+([A-Za-z0-9&\-\.\s]+?)(?:\s*$|,|\.|;)',
            r'(?i)in\s+([A-Za-z0-9&\-\.\s]+?)(?:\s*$|,|\.|;)',
        ]
        
        # Helper to clean trailing tokens accidentally captured
        def _clean_extracted_name(n: str) -> str:
            if not n:
                return n
            n = n.strip()
            # Remove trailing connectors accidentally captured
            trailing_tokens = ['for', 'into', 'in', 'gross', 'net']
            parts = n.split()
            while parts and parts[-1].lower() in trailing_tokens:
                parts.pop()
            return ' '.join(parts)

        # Extract investor name with confidence
        for pattern in name_patterns:
            match = re.search(pattern, prompt)
            if match:
                name = _clean_extracted_name(match.group(1))
                result['investor']['name'] = name
                result['investor']['confidence'] = 0.85
                break
        
        # Extract company name with confidence
        for pattern in company_patterns:
            match = re.search(pattern, prompt)
            if match:
                name = match.group(1).strip()
                result['company']['name'] = name
                result['company']['confidence'] = 0.85
                break
        
        # Check if clarification is needed
        missing_entities = []
        if not result['investor']['name'] or result['investor']['confidence'] < result['confidence_threshold']:
            missing_entities.append('investor name')
        if not result['company']['name'] or result['company']['confidence'] < result['confidence_threshold']:
            missing_entities.append('company name')
        if not result['amount']['value'] or result['amount']['confidence'] < result['confidence_threshold']:
            missing_entities.append('investment amount')
        
        if missing_entities:
            # Fallback: parse using the generic fee patterns used in can_handle
            fee_patterns = [
                r"(?i)create a fee letter for (.+?) for £?([\d,]+\.?\d*) (?:net|gross)?\s*(?:into|in) (.+)",
                r"(?i)create a letter for (.+?) for £?([\d,]+\.?\d*) (?:net|gross)?\s*(?:into|in) (.+)",
                r"(?i)generate fee letter for (.+?) investing £?([\d,]+\.?\d*) (?:net|gross)?\s*in (.+)",
                r"(?i)(.+?) £?([\d,]+\.?\d*) (.+)"
            ]
            for pat in fee_patterns:
                m = re.search(pat, prompt.strip(), re.IGNORECASE)
                if m:
                    inv, amt, comp = m.group(1).strip(), m.group(2).replace(',', ''), m.group(3).strip()
                    # Populate investor
                    if not result['investor']['name']:
                        result['investor']['name'] = inv
                        result['investor']['confidence'] = 0.9
                    # Populate amount
                    if not result['amount']['value']:
                        try:
                            result['amount']['value'] = float(amt)
                            result['amount']['confidence'] = 0.9
                        except Exception:
                            pass
                    # Populate company
                    if not result['company']['name']:
                        result['company']['name'] = comp
                        result['company']['confidence'] = 0.9
                    break

            # Recompute missing fields after fallback
            missing_entities = []
            if not result['investor']['name']:
                missing_entities.append('investor name')
            if not result['company']['name']:
                missing_entities.append('company name')
            if not result['amount']['value']:
                missing_entities.append('investment amount')

        if missing_entities:
            result['needs_clarification'] = True
            if len(missing_entities) == 1:
                result['clarification_question'] = f"Could you please clarify the {missing_entities[0]}?"
            else:
                result['clarification_question'] = f"Could you please clarify: {', '.join(missing_entities)}?"
        
        # Store extracted entities for debugging
        result['extracted_entities'] = {
            'raw_prompt': prompt,
            'net_score': net_score,
            'gross_score': gross_score,
            'missing_entities': missing_entities
        }
        
        return result
    
    def _parse_prompt(self, prompt: str):
        """Parse the fee letter prompt to extract components and investment type."""
        
        # Enhanced patterns to detect net vs gross investment
        patterns = [
            # Net investment patterns
            r"(?i)create a fee letter for (.+?) for £?([\d,]+\.?\d*) net into (.+)",
            r"(?i)create a fee letter for (.+?) to put £?([\d,]+\.?\d*) net into (.+)",
            r"(?i)generate fee letter for (.+?) investing £?([\d,]+\.?\d*) net in (.+)",
            r"(?i)(.+?) £?([\d,]+\.?\d*) net (.+)",
            
            # Gross investment patterns (existing)
            r"(?i)create a fee letter for (.+?) for £?([\d,]+\.?\d*) gross into (.+)",
            r"(?i)create a fee letter for (.+?) to put £?([\d,]+\.?\d*) gross into (.+)",
            r"(?i)generate fee letter for (.+?) investing £?([\d,]+\.?\d*) gross in (.+)",
            r"(?i)(.+?) £?([\d,]+\.?\d*) gross (.+)",
            
            # Default patterns (assume gross for backward compatibility)
            r"(?i)create a fee letter for (.+?) for £?([\d,]+\.?\d*) into (.+)",
            r"(?i)create a fee letter for (.+?) to put £?([\d,]+\.?\d*) into (.+)",
            r"(?i)create a letter for (.+?) for £?([\d,]+\.?\d*) into (.+)",
            r"(?i)create a letter for (.+?) to put £?([\d,]+\.?\d*) into (.+)",
            r"(?i)generate fee letter for (.+?) investing £?([\d,]+\.?\d*) in (.+)",
            r"(?i)(.+?) £?([\d,]+\.?\d*) (.+)"
        ]
        
        for i, pattern in enumerate(patterns):
            match = re.match(pattern, prompt.strip())
            if match:
                investor_name = match.group(1).strip()
                investment_amount = float(match.group(2).replace(",", ""))
                company_name = match.group(3).strip()
                
                # Determine if it's net or gross based on pattern index
                is_net_investment = i < 4  # First 4 patterns are net
                
                return investor_name, investment_amount, company_name, is_net_investment
        
        raise ValueError("Could not parse fee letter prompt")
    
    def _convert_html_to_text(self, html_content: str) -> str:
        """Convert HTML content to plain text for email."""
        # Simple HTML to text conversion
        import re
        
        # Remove HTML tags
        text = re.sub('<[^<]+?>', '', html_content)
        
        # Replace HTML entities
        text = text.replace('&amp;', '&')
        text = text.replace('&lt;', '<')
        text = text.replace('&gt;', '>')
        text = text.replace('&quot;', '"')
        text = text.replace('&#39;', "'")
        text = text.replace('&nbsp;', ' ')
        
        # Clean up whitespace
        text = re.sub(r'\s+', ' ', text)
        text = text.strip()
        
        return text

    @staticmethod
    def _normalize_company_name(name: str) -> str:
        """Normalize company name for matching: lowercase, trim, remove common legal suffixes and punctuation."""
        if not name:
            return ""
        s = str(name).lower().strip()
        # Remove common legal suffixes
        suffixes = [" limited", " ltd.", " ltd", " plc", " llp", " inc.", " inc", " corp.", " corp", " co.", " company"]
        for suf in suffixes:
            if s.endswith(suf):
                s = s[: -len(suf)]
                break
        # Collapse double spaces
        s = " ".join(s.split())
        return s

    def _find_company_candidates(self, input_name: str, company_records: list) -> list:
        """Find company candidates by normalized equality/startswith/contains. Returns list of matching rows."""
        target = self._normalize_company_name(input_name)
        if not target:
            return []

        scored = []
        for row in company_records:
            cname = str(row.get('Company Name', ''))
            norm = self._normalize_company_name(cname)
            score = 0
            if norm == target:
                score = 3
            elif norm.startswith(target) or target.startswith(norm):
                score = 2
            elif target in norm or norm in target:
                score = 1
            if score:
                scored.append((score, cname, row))

        # Sort by score desc then company name for stability
        scored.sort(key=lambda x: (-x[0], x[1]))
        return [r for _, _, r in scored]
    
    def _send_email(self, to_email: str, subject: str, body: str) -> bool:
        """Send email via Microsoft Graph API."""
        self.last_email_error = None
        
        try:
            # Initialize fresh mail service to use current authentication
            from microsoft_graph_mail import MicrosoftGraphMailService
            mail_service = MicrosoftGraphMailService(send_mode=self.email_mode)
            
            # Convert plain text body to HTML for better formatting
            html_body = f"<pre style='font-family: Arial, sans-serif; white-space: pre-wrap;'>{body}</pre>"
            
            # Send via Microsoft Graph
            success = mail_service.send_or_draft(
                to_address=to_email,
                subject=subject,
                html_body=html_body
            )
            
            if not success:
                self.last_email_error = "Microsoft Graph email sending failed"
                print(f"❌ Email Error: {self.last_email_error}")
            
            return success
            
        except Exception as e:
            self.last_email_error = str(e)
            print(f"❌ Microsoft Graph Email Error: {self.last_email_error}")
            return False

    def _send_email_with_cc(self, to_addresses: str, cc_address: str, subject: str, body: str) -> bool:
        """Send email via Microsoft Graph API with CC recipients."""
        self.last_email_error = None
        
        try:
            # Initialize fresh mail service to use current authentication
            from microsoft_graph_mail import MicrosoftGraphMailService
            mail_service = MicrosoftGraphMailService(send_mode=self.email_mode)
            
            # For now, send to primary recipients only (Graph API CC support would require more complex message structure)
            # Parse primary recipients
            to_list = [addr.strip() for addr in to_addresses.split(';') if addr.strip()]
            
            # Convert plain text body to HTML for better formatting
            html_body = f"<pre style='font-family: Arial, sans-serif; white-space: pre-wrap;'>{body}</pre>"
            
            # Add CC information to the body
            html_body = f"<p><strong>CC:</strong> {cc_address}</p><hr/>{html_body}"
            
            # Send to first recipient (primary)
            if to_list:
                success = mail_service.send_or_draft(
                    to_address=to_list[0],
                    subject=subject,
                    html_body=html_body
                )
                
                if not success:
                    self.last_email_error = "Microsoft Graph email sending failed"
                    print(f"❌ Email Error: {self.last_email_error}")
                
                return success
            else:
                self.last_email_error = "No valid recipients found"
                print(f"❌ Email Error: {self.last_email_error}")
                return False
            
        except Exception as e:
            self.last_email_error = str(e)
            print(f"❌ Microsoft Graph Email Error: {self.last_email_error}")
            return False

    def email_healthcheck(self) -> Dict[str, Any]:
        """Test Microsoft Graph email connectivity and authentication."""
        diagnostics: List[str] = []
        self.last_email_error = None
        
        try:
            diagnostics.append("Testing Microsoft Graph mail service...")
            
            if not self.mail_service:
                diagnostics.append("❌ Microsoft Graph mail service not initialized")
                self.last_email_error = "Mail service not available"
                return {"ok": False, "method": "microsoft_graph", "details": diagnostics, "error": self.last_email_error}
            
            # Test authentication first
            diagnostics.append("Testing Microsoft Graph authentication...")
            from microsoft_graph_auth_manager import get_auth_manager
            auth_manager = get_auth_manager()
            user_info = auth_manager.get_user_info()
            
            if not user_info:
                diagnostics.append("❌ Could not retrieve user information")
                self.last_email_error = "Authentication failed"
                return {"ok": False, "method": "microsoft_graph", "details": diagnostics, "error": self.last_email_error}
            
            user_name = user_info.get('displayName', 'Unknown')
            user_email = user_info.get('mail', 'No email')
            auth_type = user_info.get('auth_type', 'Unknown')
            diagnostics.append(f"✅ Authenticated as: {user_name} ({user_email})")
            diagnostics.append(f"🔐 Authentication type: {auth_type}")
            
            # Test email functionality (draft creation capability)
            diagnostics.append("Testing email sending capability...")
            headers = self.mail_service._get_headers()
            
            if not headers:
                diagnostics.append("❌ Could not get authorization headers for email operations")
                self.last_email_error = "Email authorization failed"
                return {"ok": False, "method": "microsoft_graph", "details": diagnostics, "error": self.last_email_error}
            
            diagnostics.append("✅ Email authorization confirmed")
            diagnostics.append(f"📧 Email mode: {self.email_mode}")
            diagnostics.append(f"📮 Target mailbox: {self.mail_service.MAILBOX}")
            diagnostics.append("✅ Ready to create drafts and send emails")
            
            return {"ok": True, "method": "microsoft_graph", "details": diagnostics}
                
        except Exception as e:
            diagnostics.append(f"❌ Microsoft Graph error: {e}")
            self.last_email_error = str(e)
            return {"ok": False, "method": "microsoft_graph", "details": diagnostics, "error": self.last_email_error}
    
    def execute_enhanced(self, prompt: str, custom_fees: Optional[Dict[str, float]] = None) -> Dict[str, Any]:
        """Enhanced fee letter generation with robust parsing, validation, fuzzy matching, and activity logging."""
        try:
            # Log the request
            if self.activity_logger:
                _ = self.activity_logger.log_activity("fee_letter_request", {
                    "prompt": prompt,
                    "timestamp": "now"
                })
            
            # Use enhanced parsing with confidence scoring
            parse_result = self.parse_request(prompt)
            
            # Check if clarification is needed
            if parse_result['needs_clarification']:
                return {
                    "success": False,
                    "message": f"🤔 {parse_result['clarification_question']}",
                    "needs_clarification": True,
                    "parse_result": parse_result,
                    "extracted_entities": parse_result['extracted_entities']
                }
            
            # Extract validated entities
            investor_name = parse_result['investor']['name']
            investment_amount = parse_result['amount']['value']
            company_name = parse_result['company']['name']
            is_net_investment = parse_result['investment_type']['type'] == 'net'
            using_default_rates = False  # Initialize default rates flag

            # Apply explicit overrides from structured UI (if provided)
            if custom_fees:
                if 'investor_name_override' in custom_fees and custom_fees['investor_name_override']:
                    investor_name = str(custom_fees['investor_name_override']).strip()
                if 'investment_amount_override' in custom_fees and custom_fees['investment_amount_override']:
                    try:
                        investment_amount = float(custom_fees['investment_amount_override'])
                    except Exception:
                        pass
                if 'company_name_override' in custom_fees and custom_fees['company_name_override']:
                    company_name = str(custom_fees['company_name_override']).strip()
                if 'investment_type_override' in custom_fees and custom_fees['investment_type_override']:
                    itype = str(custom_fees['investment_type_override']).strip().lower()
                    is_net_investment = (itype == 'net')
            
            # Log parsing confidence
            if self.activity_logger:
                self.activity_logger.log_activity("enhanced_parsing", {
                    "investor_confidence": parse_result['investor']['confidence'],
                    "company_confidence": parse_result['company']['confidence'],
                    "amount_confidence": parse_result['amount']['confidence'],
                    "investment_type_confidence": parse_result['investment_type']['confidence']
                })
            
            # Prefer local Excel adapter when available; otherwise use Google Sheets
            used_excel = False
            subscription_ref = None
            reference_preface = None
            inv_data = None
            comp_data = None
            
            # Get Excel path from config manager (preferred) or environment variable (fallback)
            excel_path = None
            try:
                # Always instantiate a fresh ConfigManager to avoid stale path
                from config_manager import ConfigManager
                excel_path = ConfigManager().get_excel_path()
            except Exception:
                pass
            
            if not excel_path:
                excel_path = os.getenv("EXCEL_PATH")
            if excel_path:
                try:
                    from adapters.excel_three_sheet_adapter import ExcelLocalThreeSheet
                    adapter = FeeLetterAgent._excel_adapter_cache.get(excel_path)
                    if adapter is None:
                        adapter = ExcelLocalThreeSheet(excel_path)
                        FeeLetterAgent._excel_adapter_cache[excel_path] = adapter
                    payload = adapter.resolve_payload(investor_name, company_name, investment_amount)
                    inv_p = payload.get("investor", {})
                    fee_ctx = payload.get("fee_context", {})
                    co_p = payload.get("company", {})
                    using_default_rates = payload.get("using_default_rates", False)
                    inv_data = {
                        'Salutation': inv_p.get('Salutation', 'Dear'),
                        'First Name': inv_p.get('First Name', ''),
                        'Last Name': inv_p.get('Last Name', ''),
                        'Contact email': inv_p.get('Contact email', ''),
                        'Investor Type': fee_ctx.get('classification', 'retail'),
                        'Custodian Client Ref': inv_p.get('Custodian Client Ref')
                    }
                    comp_data = {
                        'Company Name': co_p.get('Company Name'),
                        'Current Share Price': co_p.get('Current Share Price'),
                        'Share Class': co_p.get('Share Class') or 'Ordinary Share'
                    }
                    # Override investment type from FeeSheet context unless UI explicitly overrode it
                    ui_overrode_type = bool(custom_fees and custom_fees.get('investment_type_override'))
                    if not ui_overrode_type:
                        is_net_investment = (fee_ctx.get('investment_type') == 'net')
                    # Prepare subscription/reference
                    subscription_ref = fee_ctx.get('subscription_reference') or ''
                    reference_preface = 'CC' if inv_data['Investor Type'].lower() == 'professional' else 'CS'
                    used_excel = True
                except Exception as excel_err:
                    # Excel is the only allowed source; record the precise adapter error and do not imply a fallback
                    print(f"Excel adapter failed: {excel_err}. Excel-only mode; no fallback.")
                    # Re-raise the actual Excel error instead of the misleading message
                    raise ValueError(f"Excel adapter failed: {excel_err}")

            if not used_excel:
                # This should not happen since we're in Excel-only mode
                raise ValueError("Excel adapter failed to initialize. Check EXCEL_PATH environment variable.")

                # No company data via Google Sheets when Excel is required
                
                # Even when falling back to Google Sheets, try to overlay Subscription code from local Excel
                try:
                    excel_path_fb = os.getenv("EXCEL_PATH")
                    if excel_path_fb:
                        import pandas as _pd
                        inv_df = _pd.read_excel(excel_path_fb, sheet_name="InvestorSheet", engine="openpyxl")
                        fee_df = _pd.read_excel(excel_path_fb, sheet_name="FeeSheet", engine="openpyxl")
                        # Normalize columns
                        inv_df.columns = [str(c).strip() for c in inv_df.columns]
                        fee_df.columns = [str(c).strip() for c in fee_df.columns]
                        # Build full name
                        inv_df["__full__"] = (inv_df.get("First Name", "").astype(str).str.strip() + " " + inv_df.get("Last Name", "").astype(str).str.strip()).str.strip().str.lower()
                        full = (investor_name or "").strip().lower()
                        match = inv_df[inv_df["__full__"] == full]
                        if not match.empty:
                            cust_ref = str(match.iloc[0].get("Custodian Client Ref", "")).strip()
                            if cust_ref:
                                df = fee_df[fee_df["Custodian Client Ref"].astype(str).str.strip().str.lower() == cust_ref.lower()]
                                if not df.empty:
                                    dt_col = "Initial Deposit date/Reinvestment instruction date"
                                    if dt_col in df.columns:
                                        with _pd.option_context('mode.chained_assignment', None):
                                            df["_dt"] = _pd.to_datetime(df[dt_col], errors="coerce")
                                        df = df.sort_values(by=["_dt", "Deposit #"], ascending=[False, False], na_position="last")
                                    elif "Deposit #" in df.columns:
                                        df = df.sort_values(by=["Deposit #"], ascending=False)
                                    top = df.iloc[0]
                                    sr = str(top.get("Subscription code", "")).strip()
                                    if sr:
                                        subscription_ref = sr
                except Exception as _excel_overlay_err:
                    pass
            
            # Use proper names from data with new sheet structure
            first_name = inv_data.get('First Name', '')
            last_name = inv_data.get('Last Name', '')
            investor_name = f"{first_name} {last_name}".strip()
            company_name = comp_data['Company Name']
            
            # Validate using policy gate
            validation_result = None
            if PolicyGate:
                investment_data = {
                    "amount": investment_amount,
                    "investment_type": "net" if is_net_investment else "gross"
                }
                validation_result = PolicyGate.validate_fee_letter_request(inv_data, comp_data, investment_data)
                
                # Log validation result
                if self.activity_logger:
                    self.activity_logger.log_activity("validation_check", {
                        "investor": investor_name,
                        "company": company_name,
                        "is_valid": validation_result.is_valid,
                        "missing_fields": validation_result.missing_fields,
                        "warnings": validation_result.warnings
                    })
                
                # Return early if validation fails
                if not validation_result.is_valid:
                    return {
                        "success": False,
                        "message": f"❌ Validation Failed: {validation_result.message}",
                        "missing_fields": validation_result.missing_fields,
                        "warnings": validation_result.warnings,
                        "validation_result": validation_result
                    }
            
            # Calculate fees using enhanced calculator with custom fees and investor type logic
            fee_calculation = None
            if FeeCalculator:
                # Start from company data and enrich with Excel fee context
                calc_company_data = comp_data.copy() if comp_data else {}
                if used_excel:
                    # Pull CC Set up fee %, CC AMC %, CC Carry % from latest FeeSheet (already in payload)
                    try:
                        fee_ctx = payload.get("fee_context", {})  # type: ignore[name-defined]
                        def pct(v, default):
                            if v is None:
                                return default
                            fv = float(v)
                            return fv * 100.0 if fv <= 1.0 else fv
                        
                        calc_company_data.update({
                            'Upfront Fee %': pct(fee_ctx.get('upfront_pct'), 2.0),
                            'AMC 1–3 %': pct(fee_ctx.get('amc_pct'), 2.0),
                            'AMC 4–5 %': pct(fee_ctx.get('amc_pct'), 1.5),
                            'Perf Fee %': pct(fee_ctx.get('performance_pct'), 20.0)
                        })
                    except Exception:
                        pass
                
                # Apply UI overrides last so they take priority over Excel values
                if custom_fees:
                    try:
                        if 'upfront_fee_pct' in custom_fees and custom_fees['upfront_fee_pct'] is not None:
                            calc_company_data['Upfront Fee %'] = float(custom_fees['upfront_fee_pct'])
                        if 'amc_1_3_pct' in custom_fees and custom_fees['amc_1_3_pct'] is not None:
                            calc_company_data['AMC 1–3 %'] = float(custom_fees['amc_1_3_pct'])
                        if 'amc_4_5_pct' in custom_fees and custom_fees['amc_4_5_pct'] is not None:
                            calc_company_data['AMC 4–5 %'] = float(custom_fees['amc_4_5_pct'])
                        if 'vat_rate' in custom_fees and custom_fees['vat_rate'] is not None:
                            calc_company_data['vat_rate'] = float(custom_fees['vat_rate'])
                        # Share price override affects preview/quantity, not calculator logic directly
                        if 'share_price_override' in custom_fees and custom_fees['share_price_override'] is not None:
                            try:
                                sp_override = float(custom_fees['share_price_override'])
                                if sp_override > 0:
                                    comp_data['Current Share Price'] = sp_override
                            except Exception:
                                pass
                        # Share class override
                        if 'share_class_override' in custom_fees and custom_fees['share_class_override']:
                            comp_data['Share Class'] = str(custom_fees['share_class_override']).strip()
                    except Exception:
                        pass

                # Ensure investor type is passed for retail/professional logic
                inv_data_mod = inv_data.copy()
                if used_excel and inv_data.get('Investor Type'):
                    inv_data_mod['investor_type'] = str(inv_data.get('Investor Type')).lower()
                # Investor type override from UI
                if custom_fees and custom_fees.get('investor_type_override'):
                    inv_data_mod['investor_type'] = str(custom_fees['investor_type_override']).lower()
                # Investor type now comes from Excel Fund classification

                fee_calculation = FeeCalculator.calc_breakdown(
                    investment_amount,
                    is_net_investment,
                    calc_company_data,
                    inv_data_mod
                )
            
            # Generate fee letter using enhanced calculation or fallback
            if fee_calculation:
                # Use enhanced calculation
                gross_investment = fee_calculation.gross_investment
                total_fees = fee_calculation.total_fees
                management_fee = fee_calculation.management_fee  # used for display only
                admin_fee = fee_calculation.admin_fee  # used for display only
                total_transfer = fee_calculation.total_transfer
                
                investment_type_description = "NET" if is_net_investment else "GROSS"
                calculation_note = f"{investment_type_description} investment calculation: {fee_calculation.calculation_method}"
            else:
                # Enhanced calculation failed - return error instead of falling back to old method
                return {
                    "success": False,
                    "message": "❌ Enhanced fee calculation failed. Please check your input parameters.",
                    "error": "Enhanced calculation error"
                }
            
            # Generate the fee letter content
            # Extract share price from updated schema (Current Share Price), honoring override if provided
            try:
                sp_raw = comp_data.get('Current Share Price', 1.0)
                # handle values like "0.28", 0.28, "£0.28", "28p"
                if isinstance(sp_raw, str):
                    s = sp_raw.strip().lower()
                    s = s.replace("£", "").replace(",", "").replace(" ", "")
                    if s.endswith("p"):
                        s = s[:-1]
                    share_price = float(s)
                else:
                    share_price = float(sp_raw)
                # Explicit UI override (if any) already pushed to comp_data above; this ensures correctness
                if custom_fees and custom_fees.get('share_price_override'):
                    try:
                        share_price = float(custom_fees['share_price_override'])
                    except Exception:
                        pass
            except Exception:
                share_price = 1.0
            share_quantity = round(gross_investment / share_price, 4) if share_price else 0.0
            
            # Create fee letter preview
            preview_data = {
                "investor_name": investor_name,
                "company_name": company_name,
                "gross_investment": gross_investment,
                "total_fees": total_fees,
                "total_transfer": total_transfer,
                "share_quantity": share_quantity,
                "investment_type": investment_type_description,
                "calculation_note": calculation_note,
                "validation_warnings": [],  # No warnings - clean interface
                "input_amount": investment_amount
            }
            
            # Generate smart summary
            smart_summary = self._generate_smart_summary(
                investor_name, company_name, investment_amount, 
                investment_type_description, gross_investment, total_fees,
                validation_result, fee_calculation, using_default_rates
            )
            
            # Generate the actual fee letter content using Jinja2 template
            try:
                # Load and render the fee letter template
                template_path = "fee_letter_template.txt"
                if os.path.exists(template_path):
                    with open(template_path, 'r') as f:
                        template_content = f.read()
                    
                    template = Template(template_content)
                    
                    # Prepare template variables with CORRECT fee calculations
                    salutation = inv_data.get('Salutation', 'Dear')
                    investor_last_name = inv_data.get('Last Name', investor_name.split()[-1] if investor_name else '')
                    investment_type = investment_type_description.lower()
                    number_of_shares = f"{share_quantity:,.0f}"
                    share_type = comp_data.get('Share Class', 'Ordinary Share')
                    
                    # Display percentages with normalization: decimals (<=1) → percent
                    def pct_disp(v, default):
                        try:
                            if v is None:
                                return default
                            fv = float(v)
                            return fv * 100.0 if fv <= 1.0 else fv
                        except Exception:
                            return default
                    applied_rates = fee_calculation.breakdown.get('applied_rates', {}) if fee_calculation else {}
                    upfront_fee_pct = pct_disp(applied_rates.get('upfront_pct'), 2.0)
                    amc_1_3_pct = pct_disp(applied_rates.get('amc_1_3_pct'), 2.0)
                    amc_4_5_pct = pct_disp(applied_rates.get('amc_4_5_pct'), 1.5)
                    # Performance not in applied_rates; normalize from Excel context
                    performance_fee_pct = pct_disp((fee_ctx.get('performance_pct') if used_excel else None), 20.0)

                    # Pull values including VAT from enhanced calculation breakdown
                    fb = fee_calculation.breakdown.get('fee_breakdown', {}) if fee_calculation else {}
                    upfront_fee_value = fb.get('upfront_total', fb.get('upfront', 0.0))
                    amc_1_3_value = fb.get('amc_1_3_total', fb.get('amc_1_3', 0.0))
                    total_transfer_correct = fee_calculation.total_transfer if fee_calculation else (gross_investment + upfront_fee_value + amc_1_3_value)

                    # Format for template with proper rounding to avoid floating-point precision issues
                    upfront_fee_pct_str = f"{round(upfront_fee_pct, 2)}"
                    upfront_fee_value_str = f"{upfront_fee_value:,.2f}"
                    amc_1_3_pct_str = f"{round(amc_1_3_pct, 2)}"
                    amc_1_3_value_str = f"{amc_1_3_value:,.2f}"
                    amc_4_5_pct_str = f"{round(amc_4_5_pct, 2)}"
                    performance_fee_pct_str = f"{round(performance_fee_pct, 2)}"
                    total_transfer_str = f"{total_transfer_correct:,.2f}"
                    
                    # Build full reference for template: if Excel provided a subscription code, use it as-is.
                    # Otherwise, generate using investor type rules.
                    reference_full = None
                    if subscription_ref:
                        # If already prefixed (e.g., CC- or CS-), keep as-is; else use provided code directly
                        sr = str(subscription_ref).strip()
                        if sr.upper().startswith("CC-") or sr.upper().startswith("CS-"):
                            reference_full = sr
                        else:
                            reference_full = sr
                    else:
                        # No code from Excel; generate
                        if custom_fees and 'investor_type' in custom_fees:
                            temp_inv_data = inv_data.copy()
                            temp_inv_data['Investor Type'] = str(custom_fees['investor_type']).lower()
                            subscription_ref = self._generate_subscription_reference(temp_inv_data, comp_data)
                            reference_preface = "CC" if temp_inv_data['Investor Type'] == 'professional' else "CS"
                        else:
                            subscription_ref = self._generate_subscription_reference(inv_data, comp_data)
                            reference_preface = "CC" if inv_data.get('Investor Type', 'retail').lower() == 'professional' else "CS"
                        reference_full = f"{reference_preface}-{subscription_ref}"

                    # Hard override: if fee_ctx carries a subscription_reference from Excel, prefer that
                    try:
                        if used_excel and fee_ctx.get('subscription_reference'):
                            reference_full = str(fee_ctx.get('subscription_reference')).strip()
                    except Exception:
                        pass

                    # Expose reference_full to preview_data for UI verification
                    preview_data['reference'] = reference_full or ''
                    
                    # Render the template
                    email_content = template.render(
                        salutation=salutation,
                        investor_last_name=investor_last_name,
                        investment_type=investment_type,
                        investment_amount=f"{investment_amount:,.0f}",
                        company_name=company_name,
                        number_of_shares=number_of_shares,
                        share_type=share_type,
                        share_price=f"{share_price:.2f}",
                        upfront_fee_pct=upfront_fee_pct_str,
                        upfront_fee_value=upfront_fee_value_str,
                        amc_1_3_pct=amc_1_3_pct_str,
                        amc_1_3_value=amc_1_3_value_str,
                        amc_4_5_pct=amc_4_5_pct_str,
                        performance_fee_pct=performance_fee_pct_str,
                        total_transfer=total_transfer_str,
                         reference_full=reference_full
                    )
                else:
                    email_content = f"Template file {template_path} not found"
                    
            except Exception as template_error:
                email_content = f"Error generating template: {str(template_error)}"
            
            # Email sending is controlled by the calling function (preview vs send)
            # For now, always set to False - email will only be sent when explicitly requested
            email_sent = False
            # Resolve recipient email across possible column names (robust)
            to_email = ''
            email_candidates = []
            try:
                # 1) Preferred known keys
                preferred_keys = [
                    'Contact email',  # explicitly requested by user
                    'Contact Email', 'Email', 'Login Email', 'Contact_email',
                    'Email Address', 'Contact Email Address', 'Primary Email', 'Investor Email',
                    'email', 'e-mail'
                ]
                for key in preferred_keys:
                    if key in inv_data and inv_data.get(key):
                        to_email = str(inv_data.get(key)).strip()
                        email_candidates.append((key, to_email))
                        break
                
                # 2) If still not found, search any key containing 'email'
                if not to_email:
                    for k, v in inv_data.items():
                        if isinstance(k, str) and 'email' in k.lower() and v:
                            candidate = str(v).strip()
                            email_candidates.append((k, candidate))
                            if not to_email:
                                to_email = candidate
                
                if not to_email:
                    print("⚠️ No email address found for investor. Email-related fields:", [k for k in inv_data.keys() if isinstance(k, str) and 'email' in k.lower()])
                else:
                    print(f"✅ Resolved investor email: {to_email} (candidates: {email_candidates})")
            except Exception as e:
                print(f"⚠️ Error while resolving email: {e}")
            
            # Log successful generation
            if self.activity_logger:
                self.activity_logger.log_activity("fee_letter_generated", {
                    "investor": investor_name,
                    "company": company_name,
                    "amount": investment_amount,
                    "type": investment_type_description,
                    "gross_investment": gross_investment,
                    "total_fees": total_fees,
                    "smart_summary": smart_summary,
                    "email_sent": email_sent
                })

            # Append audit row to FeeLetterAudit sheet if possible
            try:
                # Resolve configured Excel path
                from config_manager import ConfigManager
                excel_path_cfg = ConfigManager().get_excel_path()
                # Fetch account/user info from Microsoft auth if available
                account_name = ''
                try:
                    from microsoft_graph_auth_manager import get_auth_manager
                    u = get_auth_manager().get_user_info() or {}
                    account_name = u.get('mail') or u.get('userPrincipalName') or u.get('displayName') or ''
                except Exception:
                    pass
                # Extract rates for audit (percent values)
                applied_rates = fee_calculation.breakdown.get('applied_rates', {}) if fee_calculation else {}
                def pct_disp(v, default):
                    try:
                        if v is None:
                            return default
                        fv = float(v)
                        return fv * 100.0 if fv <= 1.0 else fv
                    except Exception:
                        return default

                upfront_fee_pct = pct_disp(applied_rates.get('upfront_pct'), 2.0)
                amc_pct = pct_disp(applied_rates.get('amc_1_3_pct'), 2.0)
                carry_pct = pct_disp((fee_ctx.get('performance_pct') if used_excel else None), 20.0)

                # Build audit row
                audit_row = {
                    'Account': account_name,
                    'Custodian Client Ref': inv_data.get('Custodian Client Ref', ''),
                    'Investor Name': investor_name,
                    'Investor Email': to_email or '',
                    'Set up fee %': upfront_fee_pct,
                    'AMC %': amc_pct,
                    'Carry %': carry_pct,
                    'Amount Invested': gross_investment,
                    'Total Fees': total_fees,
                    'Gross/Net': investment_type_description,
                    'Fund': (payload.get('fee_context', {}) or {}).get('Fund', ''),
                    'Date Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                if excel_path_cfg:
                    _ = self._append_audit_row(excel_path_cfg, audit_row)
            except Exception:
                pass
            
            return {
                "success": True,
                "message": smart_summary,
                "preview_data": preview_data,
                "fee_calculation": fee_calculation,
                "validation_result": validation_result,
                "investor_data": inv_data,
                "company_data": comp_data,
                "smart_summary": smart_summary,
                "email_content": email_content,  # Add the actual fee letter content
                "email_sent": email_sent,  # Add email sending status
                "to_email": to_email  # Add email address for sending later
            }
            
        except Exception as e:
            error_msg = f"❌ Error generating fee letter: {str(e)}"
            
            # Log error
            if self.activity_logger:
                self.activity_logger.log_activity("fee_letter_error", {
                    "prompt": prompt,
                    "error": str(e)
                })
            
            return {
                "success": False,
                "message": error_msg,
                "error": str(e)
            }
    
    # OLD execute method completely removed - no more automatic email sending
    # All fee letter generation now goes through execute_enhanced method only
    
    def _generate_smart_summary(self, investor_name: str, company_name: str, 
                               investment_amount: float, investment_type: str,
                               gross_investment: float, total_fees: float,
                               validation_result: Optional[Any] = None,
                               fee_calculation: Optional[Any] = None,
                               using_default_rates: bool = False) -> str:
        """Generate single-prompt summary with key details"""
        
        # Do not display risk level/classification in summary per user preference
        classification = None
        
        # Build fee breakdown
        fee_breakdown = []
        if fee_calculation and hasattr(fee_calculation, 'breakdown'):
            breakdown = fee_calculation.breakdown
            if 'fee_breakdown' in breakdown:
                fees = breakdown['fee_breakdown']
                if fees.get('upfront', 0) > 0:
                    fee_breakdown.append(f"Upfront: £{fees['upfront']:,.2f}")
                if fees.get('amc_1_3', 0) > 0:
                    fee_breakdown.append(f"AMC 1-3: £{fees['amc_1_3']:,.2f}")
                if fees.get('amc_4_5', 0) > 0:
                    fee_breakdown.append(f"AMC 4-5: £{fees['amc_4_5']:,.2f}")
        
        if not fee_breakdown:
            fee_breakdown = [f"Total: £{total_fees:,.2f}"]
        
        # Check for issues
        issues = []
        ready_status = "Ready to send"
        
        if validation_result:
            # Suppress non-critical warnings (e.g., KYC/AML pointers) from summary
            if hasattr(validation_result, 'compliance_violations') and validation_result.compliance_violations:
                # Only surface hard compliance violations
                issues.extend(validation_result.compliance_violations[:1])
                ready_status = "Requires review"
        
        if not issues:
            issues = ["none"]
        
        # Generate comprehensive summary
        inv_line = f"**Investor:** {investor_name}"
        summary = (
            f"✅ **Fee Letter Generated**\n\n"
            f"{inv_line}\n"
            f"**Investment:** £{investment_amount:,.2f} {investment_type} → £{gross_investment:,.2f} into {company_name}\n"
            f"**Fees:** {', '.join(fee_breakdown)}\n"
            f"**Status:** {ready_status}"
        )
        
        # Add warning for default rates
        if using_default_rates:
            summary += "\n\n⚠️ **Using default rates** - No fee data found in Excel for this investor"
        
        return summary

    def _append_audit_row(self, excel_path: str, row: Dict[str, Any]) -> bool:
        """Append a single audit row to a lightweight audit workbook near the main Excel file.
        For performance, we write to '<dir>/FeeLetterAudit.xlsx' instead of the large source workbook.
        Creates the file/sheet with headers if needed.
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl import Workbook
            from pathlib import Path
        except Exception:
            return False

        columns = [
            'Account', 'Custodian Client Ref', 'Investor Name', 'Investor Email',
            'Set up fee %', 'AMC %', 'Carry %', 'Amount Invested', 'Total Fees',
            'Gross/Net', 'Fund', 'Date Generated'
        ]

        try:
            audit_path = Path(excel_path).with_name('FeeLetterAudit.xlsx')
            if audit_path.exists():
                wb = load_workbook(str(audit_path))
            else:
                wb = Workbook()
                # Ensure a clean first sheet named FeeLetterAudit
                ws0 = wb.active
                ws0.title = 'FeeLetterAudit'
            ws = wb['FeeLetterAudit'] if 'FeeLetterAudit' in wb.sheetnames else wb.create_sheet('FeeLetterAudit')
            start_row = ws.max_row + 1 if ws.max_row >= 1 else 1
            if ws.max_row < 1 or (ws.max_row == 1 and ws.cell(row=1, column=1).value is None):
                # write headers
                for idx, h in enumerate(columns, start=1):
                    ws.cell(row=1, column=idx, value=h)
                start_row = 2

            # prepare values in column order
            values = [row.get(col, '') for col in columns]
            for idx, val in enumerate(values, start=1):
                ws.cell(row=start_row, column=idx, value=val)

            wb.save(str(audit_path))
            return True
        except Exception:
            # Fallback to CSV append if Excel save fails (e.g., file locked)
            try:
                import csv
                from pathlib import Path
                csv_path = Path(excel_path).with_name('FeeLetterAudit.csv')
                file_exists = csv_path.exists()
                with open(csv_path, 'a', newline='', encoding='utf-8') as f:
                    w = csv.DictWriter(f, fieldnames=columns)
                    if not file_exists:
                        w.writeheader()
                    w.writerow({k: row.get(k, '') for k in columns})
                return True
            except Exception:
                return False
    
    def get_disambiguation_candidates(self, name: str, candidates: List[Dict[str, Any]], 
                                    name_field: str = 'name') -> List[Dict[str, Any]]:
        """Get top 3 candidates for disambiguation with key identifying fields"""
        if not FuzzyMatcher:
            return []
        
        # Get fuzzy matches with scores
        scored_candidates = []
        for candidate in candidates:
            if name_field not in candidate:
                continue
            
            score = FuzzyMatcher.similarity_score(name, candidate[name_field])
            if score >= 0.3:  # Lower threshold for disambiguation
                candidate_copy = candidate.copy()
                candidate_copy['_match_score'] = score
                scored_candidates.append(candidate_copy)
        
        # Sort by score and return top 3
        scored_candidates.sort(key=lambda x: x['_match_score'], reverse=True)
        
        # Add identifying fields for disambiguation
        for candidate in scored_candidates[:3]:
            identifying_fields = []
            for field in ['email', 'Email', 'postcode', 'Postcode', 'phone', 'Phone']:
                if field in candidate and candidate[field]:
                    identifying_fields.append(f"{field}: {candidate[field]}")
            candidate['_identifying_info'] = ', '.join(identifying_fields[:2])  # Max 2 fields
        
        return scored_candidates[:3]
    
    def _generate_subscription_reference(self, inv_data: Dict[str, Any], comp_data: Dict[str, Any]) -> str:
        """
        Generate subscription reference based on investor and fund type
        Format: PedersenJ-EIS-1 (Professional) or PedersenJ-EIS-2 (Retail)
        """
        try:
            # Use new sheet structure with First Name + Last Name
            first_name = inv_data.get('First Name', '')
            last_name = inv_data.get('Last Name', '')
            
            if first_name and last_name:
                # Format: LastnameFirstInitial (e.g., PedersenJ)
                first_initial = first_name[0] if first_name else ''
                name_prefix = f"{last_name}{first_initial}"
            else:
                # Fallback if names are missing
                name_prefix = f"{first_name}{last_name}".replace(' ', '')[:8] if f"{first_name}{last_name}".strip() else 'INVESTOR'
            
            fund_type = comp_data.get('Fund Type', 'EIS')
            investor_type = inv_data.get('Investor Type', 'retail').lower()
            
            if 'knowledge-intensive' in fund_type.lower() or 'kic' in fund_type.lower():
                fund_code = 'KIC'
            elif 'eis' in fund_type.lower():
                fund_code = 'EIS'
            else:
                fund_code = 'EIS'
            
            if investor_type == 'professional':
                suffix = '1'
            else:
                suffix = '2'
            
            return f"{name_prefix}-{fund_code}-{suffix}"
            
        except Exception as e:
            return f"REF-EIS-1"  # Fallback reference
    
    def send_fee_letter_email(self, to_email: str, company_name: str, email_content: str) -> bool:
        """
        Send fee letter email when explicitly requested (not during preview)
        """
        try:
            subject = f"Fee Letter Confirmation – {company_name}"
            # Local log to trace UI-triggered send attempts
            try:
                with open("email_send_log.txt", "a", encoding="utf-8") as f:
                    f.write(f"SEND_ATTEMPT|to={to_email}|subject={subject}|len={len(email_content)}\n")
            except Exception:
                pass
            email_sent = self._send_email(to_email, subject, email_content)
            if email_sent:
                print(f"📧 Fee letter email sent successfully to {to_email}")
                try:
                    with open("email_send_log.txt", "a", encoding="utf-8") as f:
                        f.write(f"SEND_SUCCESS|to={to_email}\n")
                except Exception:
                    pass
            else:
                print(f"❌ Failed to send fee letter email to {to_email}")
                try:
                    with open("email_send_log.txt", "a", encoding="utf-8") as f:
                        f.write(f"SEND_FAIL|to={to_email}|err={self.last_email_error}\n")
                except Exception:
                    pass
            return email_sent
        except Exception as e:
            print(f"❌ Error sending fee letter email: {e}")
            try:
                with open("email_send_log.txt", "a", encoding="utf-8") as f:
                    f.write(f"SEND_EXCEPTION|to={to_email}|err={e}\n")
            except Exception:
                pass
            return False
    
    def send_team_review_email(self, primary_recipients: list, cc_recipient: str, company_name: str, email_content: str, investor_name: str) -> bool:
        """Send fee letter to team for review before investor delivery"""
        try:
            subject = f"TEAM REVIEW - Fee Letter for {investor_name} - {company_name}"
            
            # Join primary recipients with semicolon for TO field
            to_addresses = "; ".join(primary_recipients)
            
            # Send email with CC
            return self._send_email_with_cc(to_addresses, cc_recipient, subject, email_content)
        except Exception as e:
            self.last_email_error = str(e)
            return False
    
    def send_team_review_email_no_cc(self, primary_recipients: list, company_name: str, email_content: str, investor_name: str) -> bool:
        """Send fee letter to team for review without CC recipient"""
        try:
            subject = f"TEAM REVIEW - Fee Letter for {investor_name} - {company_name}"
            
            # Join primary recipients with semicolon for TO field
            to_addresses = "; ".join(primary_recipients)
            
            # Send email without CC (use regular send method)
            return self._send_email(to_addresses, subject, email_content)
        except Exception as e:
            self.last_email_error = str(e)
            return False
    
    def _get_investor_data_by_name(self, full_name: str) -> Dict[str, Any]:
        """
        Get investor data from the updated sheet structure using First Name + Last Name
        """
        try:
            gc = gspread.authorize(self.sheets_creds)
            spreadsheet = gc.open(self.spreadsheet_name)
            investor_sheet = spreadsheet.worksheet("InvestorData")
            
            # Get all data from the sheet
            all_data = investor_sheet.get_all_records()
            
            # Parse the input name (case-insensitive, ignore trailing connectors like 'for')
            clean = full_name.strip()
            for trailing in [' for', ' into', ' in']:
                if clean.lower().endswith(trailing):
                    clean = clean[: -len(trailing)].strip()
                    break
            name_parts = clean.split()
            if len(name_parts) < 2:
                # If only one name provided, try to match against first or last name
                search_name = name_parts[0].lower()
                for row in all_data:
                    first_name = str(row.get('First Name', '')).lower()
                    last_name = str(row.get('Last Name', '')).lower()
                    if search_name in first_name or search_name in last_name:
                        return row
            else:
                # Try to match first and last name
                search_first = name_parts[0].lower()
                search_last = name_parts[-1].lower()
                
                # First try exact match
                for row in all_data:
                    first_name = str(row.get('First Name', '')).lower()
                    last_name = str(row.get('Last Name', '')).lower()
                    if first_name == search_first and last_name == search_last:
                        return row
                
                # Then try partial match
                for row in all_data:
                    first_name = str(row.get('First Name', '')).lower()
                    last_name = str(row.get('Last Name', '')).lower()
                    if (search_first in first_name or first_name in search_first) and \
                       (search_last in last_name or last_name in search_last):
                        return row
            
            # If no match found, raise an error with available names
            available_names = []
            for row in all_data[:5]:  # Show first 5 for reference
                first = row.get('First Name', '')
                last = row.get('Last Name', '')
                if first or last:
                    available_names.append(f"{first} {last}".strip())
            
            raise ValueError(f"Investor '{full_name}' not found. Available investors: {', '.join(available_names)}")
            
        except Exception as e:
            raise ValueError(f"Error finding investor data: {str(e)}")
    
    def refresh_excel_cache(self) -> bool:
        """Clear the Excel adapter cache to force fresh data loading."""
        try:
            # Clear the class-level cache that stores adapter instances
            if hasattr(FeeLetterAgent, '_excel_adapter_cache'):
                FeeLetterAgent._excel_adapter_cache.clear()
            
            # Clear any instance-level cached adapter
            if hasattr(self, '_excel_adapter'):
                delattr(self, '_excel_adapter')
            
            # Force re-initialization on next use
            self._excel_adapter = None
            
            return True
        except Exception as e:
            raise ValueError(f"Error refreshing Excel cache: {str(e)}")
    
    def debug_company_search(self) -> Dict[str, Any]:
        """Debug company search by listing available companies in Excel."""
        try:
            from adapters.excel_three_sheet_adapter import ExcelLocalThreeSheet
            import os, time
            
            # Get Excel file path from config manager
            try:
                from config_manager import ConfigManager
                config_manager = ConfigManager()
                excel_file_path = config_manager.get_excel_path()
            except Exception as e:
                excel_file_path = None
            
            if not excel_file_path:
                # Fallback: try environment variable
                import os
                excel_file_path = os.getenv('EXCEL_PATH')
            
            if not excel_file_path:
                return {
                    "total_companies": 0,
                    "company_names": [],
                    "error": "No Excel file path configured. Please go to Settings to configure your Excel file."
                }
            
            # Capture file mtime for verification
            try:
                mtime = os.path.getmtime(excel_file_path)
                excel_mtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(mtime))
            except Exception:
                excel_mtime = None

            # Initialize adapter and get company data (fresh read)
            adapter = ExcelLocalThreeSheet(excel_file_path)
            
            # Get company data from the adapter
            if hasattr(adapter, 'cos') and adapter.cos is not None:
                company_data = adapter.cos
                
                # Try common column names for company names
                name_columns = ['Company Name', 'Name', 'Company', 'Legal Name']
                company_names = []
                
                for col in name_columns:
                    if col in company_data.columns:
                        names = company_data[col].dropna().astype(str).tolist()
                        company_names.extend([name for name in names if name.strip()])
                        break
                
                # If no standard column found, try all columns that might contain company names
                if not company_names:
                    for col in company_data.columns:
                        if any(word in col.lower() for word in ['company', 'name', 'firm', 'business']):
                            names = company_data[col].dropna().astype(str).tolist()
                            company_names.extend([name for name in names if name.strip()])
                            break
                
                unique_companies = sorted(list(set(company_names)))
                return {
                    "total_companies": len(unique_companies),
                    "company_names": unique_companies,
                    "excel_file": excel_file_path,
                    "excel_mtime": excel_mtime,
                    "columns_found": list(company_data.columns),
                    "contains_dingus": any("dingus" in name.lower() for name in unique_companies),
                    "contains_pingus": any("pingus" in name.lower() for name in unique_companies)
                }
            else:
                return {
                    "total_companies": 0,
                    "company_names": [],
                    "error": "Could not access company data from Excel adapter",
                    "excel_file": excel_file_path,
                    "excel_mtime": excel_mtime
                }
                
        except Exception as e:
            return {
                "total_companies": 0,
                "company_names": [],
                "error": f"Error debugging company search: {str(e)}"
            }

    def list_companies(self) -> Dict[str, Any]:
        """Return companies with key details for UI listing."""
        try:
            from adapters.excel_three_sheet_adapter import ExcelLocalThreeSheet
            try:
                from config_manager import ConfigManager
                xp = ConfigManager().get_excel_path()
            except Exception:
                xp = os.getenv('EXCEL_PATH')
            if not xp:
                return {"total": 0, "rows": [], "excel_file": None}
            # Reuse cached adapter to avoid re-reading Excel on each keystroke
            adapter = FeeLetterAgent._excel_adapter_cache.get(xp)
            if adapter is None:
                adapter = ExcelLocalThreeSheet(xp)
                FeeLetterAgent._excel_adapter_cache[xp] = adapter
            df = getattr(adapter, 'cos', None)
            if df is None:
                return {"total": 0, "rows": [], "excel_file": xp}
            cols = [c for c in ['Company Name', 'Current Share Price', 'Share Class', 'Company Number'] if c in df.columns]
            rows = []
            for _, r in df.iterrows():
                item = {c: r.get(c) for c in cols}
                rows.append(item)
            return {"total": len(rows), "rows": rows, "excel_file": xp}
        except Exception as e:
            return {"total": 0, "rows": [], "error": str(e)}

    def list_investors(self, query: Optional[str] = None) -> Dict[str, Any]:
        """Return investors with key details, fee columns, and missing-data flags.
        Supports optional case-insensitive name search.
        """
        try:
            from adapters.excel_three_sheet_adapter import ExcelLocalThreeSheet
            try:
                from config_manager import ConfigManager
                xp = ConfigManager().get_excel_path()
            except Exception:
                xp = os.getenv('EXCEL_PATH')
            if not xp:
                return {"total": 0, "rows": [], "excel_file": None}
            # Use cached adapter when possible
            adapter = FeeLetterAgent._excel_adapter_cache.get(xp)
            if adapter is None:
                adapter = ExcelLocalThreeSheet(xp)
                FeeLetterAgent._excel_adapter_cache[xp] = adapter
            inv_df = getattr(adapter, 'inv', None)
            fees_df = getattr(adapter, 'fees', None)
            if inv_df is None:
                return {"total": 0, "rows": [], "excel_file": xp}
            df = inv_df.copy()
            if 'First Name' in df.columns and 'Last Name' in df.columns:
                df['__FullName'] = (df['First Name'].astype(str).str.strip() + ' ' + df['Last Name'].astype(str).str.strip()).str.strip()
            else:
                df['__FullName'] = df.get('Investor Name', '')
            if query:
                q = str(query).strip().lower()
                df = df[df['__FullName'].astype(str).str.lower().str.contains(q, na=False)]
            # Required fields mapped to the keys we actually output in `row`
            required = ['First Name', 'Last Name', 'Investor Email', 'Custodian Client Ref', 'Set up fee %', 'AMC %', 'Carry %', 'Gross/Net']
            out = []
            # Build fee lookup by Custodian Client Ref for fallback
            fee_lookup = {}
            if fees_df is not None and 'Custodian Client Ref' in fees_df.columns:
                try:
                    df_fees = fees_df.copy()
                    for _, fr in df_fees.iterrows():
                        key = str(fr.get('Custodian Client Ref', '')).strip()
                        if key and key not in fee_lookup:
                            fee_lookup[key] = fr
                except Exception:
                    pass
            def norm_pct(v):
                try:
                    f = float(v)
                    return f*100.0 if f <= 1.0 else f
                except Exception:
                    return None
            for _, r in df.iterrows():
                full_name = r.get('__FullName')
                ref = str(r.get('Custodian Client Ref', '')).strip()
                
                # Check what's actually missing from the Excel data BEFORE creating fallbacks
                missing_from_excel = []
                
                # Check investor sheet fields
                if not str(r.get('First Name', '')).strip():
                    missing_from_excel.append('First Name')
                if not str(r.get('Last Name', '')).strip():
                    missing_from_excel.append('Last Name')
                if not str(r.get('Contact email', '')).strip() and not str(r.get('Email', '')).strip():
                    missing_from_excel.append('Investor Email')
                if not ref:
                    missing_from_excel.append('Custodian Client Ref')
                
                # Check fee sheet fields
                fee_row = None
                try:
                    if hasattr(adapter, 'find_fee_row') and ref:
                        fee_row = adapter.find_fee_row(ref, None, investor_full_name=full_name)
                except Exception:
                    fee_row = None
                if fee_row is None:
                    fee_row = fee_lookup.get(ref)
                
                if fee_row is not None:
                    if not str(fee_row.get('CC Set up fee %', '')).strip():
                        missing_from_excel.append('Set up fee %')
                    if not str(fee_row.get('CC AMC %', '')).strip():
                        missing_from_excel.append('AMC %')
                    if not str(fee_row.get('CC Carry %', '')).strip():
                        missing_from_excel.append('Carry %')
                    if not str(fee_row.get('Gross/Net', '')).strip():
                        missing_from_excel.append('Gross/Net')
                else:
                    # If no fee row found, all fee fields are missing
                    missing_from_excel.extend(['Set up fee %', 'AMC %', 'Carry %', 'Gross/Net'])
                
                row = {
                    'Full Name': full_name,
                    'First Name': r.get('First Name'),
                    'Last Name': r.get('Last Name'),
                    'Investor Email': r.get('Contact email') or r.get('Email'),
                    'Custodian Client Ref': ref,
                }
                
                if fee_row is not None:
                    row['Set up fee %'] = norm_pct(fee_row.get('CC Set up fee %'))
                    row['AMC %'] = norm_pct(fee_row.get('CC AMC %'))
                    row['Carry %'] = norm_pct(fee_row.get('CC Carry %'))
                    row['Gross/Net'] = fee_row.get('Gross/Net')
                    fund_val = fee_row.get('Fund')
                    row['Investor Type'] = 'Professional' if isinstance(fund_val, str) and 'pro' in fund_val.lower() else 'Retail'
                else:
                    row['Set up fee %'] = None
                    row['AMC %'] = None
                    row['Carry %'] = None
                    row['Gross/Net'] = None
                    row['Investor Type'] = None
                
                # Use the actual missing fields from Excel, not the processed row
                row['missing'] = missing_from_excel
                out.append(row)
            return {"total": len(out), "rows": out, "excel_file": xp}
        except Exception as e:
            return {"total": 0, "rows": [], "error": str(e)}